import openpyxl


def get_row_count(file, sheet_name):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheet_name)
    return sheet.max_row


def get_col_count(file, sheet_name):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheet_name)
    return sheet.max_column


def read_data(file, sheet_name, row_num, col_num):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheet_name)
    return sheet.cell(row_num, col_num).value


def write_data(file, sheet_name, row_num, col_num, data):
    wb = openpyxl.load_workbook(file)
    sheet = wb.get_sheet_by_name(sheet_name)
    sheet.cell(row_num, col_num).value = data
    wb.save(file)
