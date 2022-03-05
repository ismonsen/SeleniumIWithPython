import openpyxl

path = 'data1.xlsx'

wb = openpyxl.load_workbook(path)
sheet = wb['Sheet1']

rows = sheet.max_row
cols = sheet.max_column

print(rows)
print(cols)

for row in range(1, rows + 1):
    for col in range(1, cols + 1):
        print(sheet.cell(row=row, column=col).value, end='\t\t')
    print()
